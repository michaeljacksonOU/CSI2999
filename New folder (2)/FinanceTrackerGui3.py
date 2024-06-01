import tkinter as tk
from tkinter import ttk
import openpyxl

def budget_window():
    # Create the budget window
    global budget_window
    budget_window = tk.Tk()
    budget_window.title("Monthly Budget")
    
    # Set the window to open in the center of the user's screen
    width = 300
    height = 100
    screen_width = budget_window.winfo_screenwidth()
    screen_height = budget_window.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    budget_window.geometry("{}x{}+{}+{}".format(width, height, int(x), int(y)))
    
    # Create a label telling the user to enter their budget 
    label1 = tk.Label(budget_window, text="Enter your monthly budget:")
    label1.pack()
    
    # Create a text field for the user to enter their budget
    budget_amount = tk.Entry(budget_window)
    budget_amount.pack()
    
    # Create a Confirm button
    # When the Confirm button is clicked, the close_budget_window() method is called and the budget is sent to that method
    confirm_button = tk.Button(budget_window, text="Confirm", command=lambda: close_budget_window(budget_amount.get()))
    confirm_button.pack(padx=10, pady=10)

    # Start the budget event loop
    budget_window.mainloop()


# The budget is received as a parameter from the budget_window method
def close_budget_window(budget):
    # Close the budget window
    budget_window.destroy()

    # Open the main window and send the budget to that method
    main_window(budget)


def main_window(budget):
    # Create the main window
    root = tk.Tk()
    root.title("Budget Tracker")
    
    # Set the window size to cover the entire screen without covering the taskbar
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"{screen_width}x{screen_height-100}+0+0")

    # Top Right Box
    # Create four frames
    frame1 = tk.Frame(root, bg="lightgrey", width=250, height=250)
    frame2 = tk.Frame(root, bg="lightgrey", width=250, height=250)
    frame3 = tk.Frame(root, bg="lightgrey", width=250, height=250)
    frame4 = tk.Frame(root, bg="lightgrey", width=250, height=250)

    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    frame1.grid_propagate(False)
    frame2.grid_propagate(False)
    frame3.grid_propagate(False)
    frame4.grid_propagate(False)

    # Top Left Box
    # Convert the budget amount to a float
    budget_amount = float(budget)

    # Create Budget and Funds Remaining labels
    label1 = tk.Label(frame1, text="Budget:", anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    label1.grid(row=0, column=0, padx=25, pady=25, sticky="w")
    budget_amount_label = tk.Label(frame1, text="${:,.2f}".format(budget_amount), anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    budget_amount_label.grid(row=0, column=1, padx=25, pady=25, sticky="w")
    label2 = tk.Label(frame1, text="Funds Remaining:", anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    label2.grid(row=1, column=0, padx=25, pady=25, sticky="w")
    funds_remaining_label = tk.Label(frame1, text="${:,.2f}".format(budget_amount), anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    funds_remaining_label.grid(row=1, column=1, padx=25, pady=25, sticky="w")

    # Top Right Box
    # Create the treeview to display the user's expenses
    tree = ttk.Treeview(frame2, show="headings")

    tree["columns"] = ("Name", "Price", "Category", "Priority", "Date")

    tree.column("Name", width=100, anchor="w", stretch=tk.YES)
    tree.column("Price", width=100, anchor="w", stretch=tk.YES)
    tree.column("Category", width=100, anchor="w", stretch=tk.YES)
    tree.column("Priority", width=100, anchor="w", stretch=tk.YES)
    tree.column("Date", width=100, anchor="w", stretch=tk.YES)

    tree.heading("Name", text="Name", anchor="w")
    tree.heading("Price", text="Price", anchor="w")
    tree.heading("Category", text="Category", anchor="w")
    tree.heading("Priority", text="Priority", anchor="w")
    tree.heading("Date", text="Date", anchor="w")

    tree.grid(row=0, column=0, sticky="nsew")

    frame2.grid_rowconfigure(0, weight=1)
    frame2.grid_columnconfigure(0, weight=1)
    frame2.grid_columnconfigure(1, weight=0)

    # Place the frames in a 2x2 grid
    frame1.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    frame2.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    frame3.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
    frame4.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")

    # Below Boxes
    # Create an Add New Expense button that will open a pop-up window for the user to enter/upload ther expense
    add_new_expense_button = tk.Button(root, text="Add New Expense", command=lambda: expense_popup(tree))
    add_new_expense_button.grid(row=2, column=0, columnspan=2, pady=10)
    
    # Bottom Left Box
    # Create priority level labels
    high_priority_label = tk.Label(frame3, text="High priority expenses : $0", anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    high_priority_label.grid(row=0, column=0, padx=25, pady=25, sticky="w")
    
    medium_priority_label = tk.Label(frame3, text="Medium priority expenses : $0", anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    medium_priority_label.grid(row=1, column=0, padx=25, pady=25, sticky="w")
    
    low_priority_label = tk.Label(frame3, text="Low priority expenses : $0", anchor="w", font=("Arial", 28), bg="lightgrey", fg="black")
    low_priority_label.grid(row=2, column=0, padx=25, pady=25, sticky="w")
    
    # Bottom Right Box
    # Create priority level tabs
    style = ttk.Style()
    style.configure("TNotebook.Tab", font=('Helvetica', 25,'bold'), padding=[40,0])
    notebook = ttk.Notebook(frame4, style="TNotebook")
    
    tab1 = ttk.Frame(notebook)
    tab2 = ttk.Frame(notebook)
    tab3 = ttk.Frame(notebook)
    
    notebook.add(tab1, text="High")
    notebook.add(tab2, text="Medium")
    notebook.add(tab3, text="Low")
    notebook.pack(expand=1, fill='both')

    # Call to method load_expenses()
    load_expenses(tree)

    # Start the main event loop
    root.mainloop()


def expense_popup(tree):
    # Create a pop-up window asking the user how they would like to input their expense
    expense_popup = tk.Toplevel()
    expense_popup.title("Add New Expense")
    
    # Ensure the window opens in the center of the user's screen
    width = 300
    height = 100
    screen_width = expense_popup.winfo_screenwidth()
    screen_height = expense_popup.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    expense_popup.geometry("{}x{}+{}+{}".format(width, height, int(x), int(y)))

    # Create a label asking the user how they would like to input their expense
    label1 = tk.Label(expense_popup, text="How would you like to input your expense?")
    label1.pack(pady=5)

    # Create a frame to contain the buttons
    button_frame = tk.Frame(expense_popup)
    button_frame.pack()

    # Create a Enter Manually button to allow the user to input their expense
    manual_button = tk.Button(expense_popup, text="Enter Manually", command=lambda: enter_manually_popup(tree))
    manual_button.pack(side=tk.LEFT, padx=30)

    # Create an Upload Receipt button to allow the user to upload 
    # their receipt from their computer files
    upload_button = tk.Button(expense_popup, text="Upload Receipt")
    upload_button.pack(side=tk.RIGHT, padx=30)
 

def enter_manually_popup(tree):
    # Create a pop-up window for the user to manually enter their expense
    enter_manually_popup = tk.Toplevel()
    enter_manually_popup.title("Manual Expense Entry")
    
    # Ensure the window opens in the center of the user's screen
    width = 800
    height = 250
    screen_width = enter_manually_popup.winfo_screenwidth()
    screen_height = enter_manually_popup.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    enter_manually_popup.geometry("{}x{}+{}+{}".format(width, height, int(x), int(y)))

    # Create name, price, category, date, and priority labels
    name_label = tk.Label(enter_manually_popup, text="Name")
    name_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    price_label = tk.Label(enter_manually_popup, text="Price")
    price_label.grid(row=0, column=1, padx=10, pady=10, sticky="w")

    category_label = tk.Label(enter_manually_popup, text="Category")
    category_label.grid(row=0, column=3, padx=10, pady=10, sticky="w")

    priority_label = tk.Label(enter_manually_popup, text="Priority")
    priority_label.grid(row=0, column=4, padx=10, pady=10, sticky="w")

    date_label = tk.Label(enter_manually_popup, text="Date")
    date_label.grid(row=0, column=5, padx=10, pady=10, sticky="w")

    # Create the entry text fields
    global name_entry
    name_entry = tk.Entry(enter_manually_popup)
    name_entry.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

    global price_entry
    price_entry = tk.Entry(enter_manually_popup)
    price_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

    # Create a combobox for category selection
    category_options = ["Food", "Personal", "Work", "Home", "Misc"]
    category_var = tk.StringVar(enter_manually_popup)
    global category_combobox
    category_combobox = ttk.Combobox(enter_manually_popup, textvariable=category_var, values=category_options, state="readonly")
    category_combobox.grid(row=1, column=3, padx=10, pady=10, sticky="ew")
    category_combobox.current(0)

    # Create  a combobox for priority selection
    priority_options = ["High", "Medium", "Low"]
    priority_var = tk.StringVar(enter_manually_popup)
    global priority_combobox
    priority_combobox = ttk.Combobox(enter_manually_popup, textvariable=priority_var, values=priority_options, state="readonly")
    priority_combobox.grid(row=1, column=4, padx=10, pady=10, sticky="ew")
    priority_combobox.current(0)

    global date_entry
    date_entry = tk.Entry(enter_manually_popup)
    date_entry.grid(row=1, column=5, padx=10, pady=10, sticky="ew")

    # Create a Confirm button will
    confirm_button = tk.Button(enter_manually_popup, text="Confirm", command=lambda: get_expense_info(tree))
    confirm_button.grid(row=3, column=1, columnspan=4, padx=10, pady=10, sticky="nsew")


def load_expenses(tree):
    # Load the the excel file from the filepath and access the active sheet
    filepath = r"C:\Users\User\Documents\New folder (2)\Expenses.xlsx"
    workbook = openpyxl.load_workbook(filepath)
    sheet1 = workbook.active

    # Load the expenses data from the active sheet of the excel file into the Treeview
    # Start the iteration from row 2 of the sheet and yield cell values only 
    # (rather than cell objects in order to save memory since we do not need additional cell properties)  
    for current_row in sheet1.iter_rows(min_row=2, values_only=True):
        # Parent argument is empty: "" 
        # Enter expense data at the end of the Treeview: "end" 
        # Insert expense data extracted from current row of the sheet: data=current_row
        tree.insert("", "end", values=current_row)


# Method get_expense_info() modified from "Code First with Hala" YouTube video:
# https://www.youtube.com/watch?v=8m4uDS_nyCk&pp=ygUlcHl0aG9uIGV4Y2VsIGFwcCBjb2RlIGZpcnN0IHdpdGggaGFsYQ%3D%3D
def get_expense_info(tree):
    filepath = r"C:\Users\User\Documents\New folder (2)\Expenses.xlsx"
    workbook = openpyxl.load_workbook(filepath)
    sheet1 = workbook.active
    expense_info = [name_entry.get(), price_entry.get(), category_combobox.get(), priority_combobox.get(), date_entry.get()]
    sheet1.append(expense_info)
    workbook.save(filepath)

    # Insert expense info into the treeview 
    tree.insert('', tk.END, values=expense_info)


# Call to method budget_window() that will start up the program
budget_window()