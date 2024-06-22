import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import openpyxl
import pandas as pd
import os
from customtkinter.windows.widgets import ctk_label
from tkcalendar import DateEntry
import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.font_manager import FontProperties
from PIL import Image, ImageTk
import pytesseract
import json
from openai import OpenAI
from dotenv import load_dotenv

pytesseract.pytesseract.tesseract_cmd=r"C:\Users\User\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

# Load .env
load_dotenv()

# API-key
#api_key = os.getenv('OPENAI_API_KEY')

#client = OpenAI(api_key=api_key)

# Note: assignment of the file path to expense_file must be changed accordingly to the unique location where is saved on each individual user's end system 
expense_file = r"C:\Users\User\Documents\June22Test\expenses.xlsx"

funds_remaining = 0.0

tab4 = None
food_label = None
personal_label = None
work_label = None
home_label = None
transportation_label = None
recurring_label = None
misc_label = None

# Set app mode and theme
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
mode = "dark"

# Method to switch the app mode between light and dark mode
def change_theme():
    global mode
    if mode == "dark":
        ctk.set_appearance_mode("light")
        mode = "light"
        apply_light_mode_style()
        apply_light_mode_tabs()
        apply_theme_to_labels_in_tab4("light")

    else:
        ctk.set_appearance_mode("dark")
        mode = "dark"
        apply_dark_mode_style()
        apply_dark_mode_tabs()
        apply_theme_to_labels_in_tab4("dark")


def welcome_window():
    global root
    welcome = ctk.CTkToplevel(root)
    welcome.title("Welcome to ExpenseExpert")

    welcome_width = 400
    welcome_height = 300

    screen_width = welcome.winfo_screenwidth()
    screen_height = welcome.winfo_screenheight()

    # Calculate the x and y coordinates for the window to be centered
    x = (screen_width // 2) - (welcome_width // 2)
    y = (screen_height // 2) - (welcome_height // 2)

    # Set the geometry of the window
    welcome.geometry(f'{welcome_width}x{welcome_height}+{x}+{y}')

    if mode == "dark":
        apply_dark_mode_style_welcome()
    else:
        apply_light_mode_style_welcome()

    # Create and place the welcome label
    welcome_label = ctk.CTkLabel(welcome, text="Welcome to ExpenseExpert!", font=("Arial", 16))
    welcome_label.pack(pady=20)

    # Load and resize the question mark icon
    question_icon = Image.open(r"C:\Users\User\Documents\June22Test\question_mark (1).png")
    question_icon = question_icon.resize((20, 20), Image.Resampling.LANCZOS)  # Resize to smaller size
    question_icon = ctk.CTkImage(question_icon)

    # Create a frame to hold the question mark button
    frame = ctk.CTkFrame(welcome)
    frame.pack(fill="both", expand=True)

    # Create and place the question mark button in the top left corner
    button_size = 35
    question_button = ctk.CTkButton(frame, width=button_size, height=button_size, image=question_icon, text="", command=show_instructions)
    question_button.image = question_icon  # Keep a reference to the image
    question_button.place(x=10, y=10)  # Adjust x, y to position the icon as needed

    # Create and place the close button at the bottom
    close_button = ctk.CTkButton(welcome, text="Get Started", command=lambda: close_welcome(welcome))
    close_button.pack(side="bottom")


def close_welcome(welcome):
    welcome.destroy()
    root.deiconify()  # Show the main root window


# Function to show instructions
def show_instructions():
    instructions = (
        "Instructions on how to use the ExpenseExpert app:\n\n"
        "1. To add a new expense, click 'Add New Expense' and fill in the details.\n"
        "2. To delete an expense, select it from the list and click 'Delete Expense'.\n"
        "3. To upload a receipt, click 'Upload Receipt' and follow the instructions.\n"
        "4. Switch between light and dark modes using the toggle button.\n"
    )
    messagebox.showinfo("Instructions", instructions)


# Method to display main window elements:
# Frame 1: budget and funds remaining
# Frame 2: treeview of all the user's entered expenses
# Frame 3: pie chart of distribution of entered expenses by category
# Frame 4: treeview of entered expenses sorted by priority tabs 
def main_window():
    global root, tree, tree1, tree2, tree3, funds_remaining_label, tab4, tab5
    global food_label, personal_label, work_label, home_label, transportation_label, recurring_label, misc_label

    # Create the main window 
    root = ctk.CTk()
    root.withdraw()
    root.title("ExpenseExpert")

    window_width = 1280
    window_height = 720

    # Get the screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calculate the x and y coordinates for the window to be centered
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    # Set the geometry of the window
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')

    # Create and place the frames in the grid
    frame1_width, frame1_height = 620, 340
    frame2_width, frame2_height = 620, 340
    frame3_width, frame3_height = 620, 340
    frame4_width, frame4_height = 620, 340
    frame5_width, frame5_height = 620, 40

    # Create and place the frames in the grid
    frame1 = ctk.CTkFrame(root, width=frame1_width, height=frame1_height)
    frame2 = ctk.CTkFrame(root, width=frame2_width, height=frame2_height)
    frame3 = ctk.CTkFrame(root, width=frame3_width, height=frame3_height)
    frame4 = ctk.CTkFrame(root, width=frame4_width, height=frame4_height)
    frame5 = ctk.CTkFrame(root, width=frame5_width, height=frame5_height)

    frame1.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    frame2.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    frame3.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
    frame4.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
    frame5.grid(row=2, column=0, padx=10, pady=10, sticky="w")

    # Set propagation to false for all frames in order to prevent them from resizing based on the size of the children widgets within them
    frame1.propagate(False)
    frame2.propagate(False)
    frame3.propagate(False)
    frame4.propagate(False)
    frame5.propagate(False)

    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    # Create a toggle bar to allow the user to switch between light and dark mode
    switch_var = ctk.StringVar(value="on")
    switch = ctk.CTkSwitch(frame5, text="Mode", command=change_theme, variable=switch_var, onvalue="on", offvalue="off")
    switch.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    # Frame 2 (Top Right)
    if mode == "dark":
        apply_dark_mode_style()

    else:
        apply_light_mode_style()

    # Create a treeview to display the user's entered expenses
    tree = ttk.Treeview(frame2, show="headings", style="Treeview")
    tree["columns"] = ("Name", "Type", "Price", "Priority", "Date")

    for col in tree["columns"]:
        tree.column(col, width=100, anchor="w", stretch=tk.YES)
        tree.heading(col, text=col, anchor="w")

    scrollbar = ttk.Scrollbar(frame2, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=1, sticky='ns')

    tree.configure(yscrollcommand=scrollbar.set)
    tree.grid(row=0, column=0, sticky="nsew")

    frame2.grid_rowconfigure(0, weight=1)
    frame2.grid_columnconfigure(0, weight=1)

    # Frame 4 (Bottom Right)
    style = ttk.Style()
    style.configure("TNotebook.Tab", font=('Helvetica', 25, 'bold'), padding=[40, 0])
    notebook1 = ttk.Notebook(frame4, style="TNotebook")

    # Create a tab containg a treeview filtered to display only High priority expenses
    tab1 = ttk.Frame(notebook1)
    tree1 = ttk.Treeview(tab1, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree1.column(col, width=90, anchor="w", stretch=tk.YES)
        tree1.heading(col, text=col)
    scrollbar1 = ttk.Scrollbar(tab1, orient="vertical", command=tree1.yview)
    tree1.configure(yscrollcommand=scrollbar1.set)
    tree1.pack(side="left", fill="both", expand=True)
    scrollbar1.pack(side="right", fill="y")

    # Create a tab containg a treeview filtered to display only Medium priority expenses
    tab2 = ttk.Frame(notebook1)
    tree2 = ttk.Treeview(tab2, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree2.column(col, width=90, anchor="w", stretch=tk.YES)
        tree2.heading(col, text=col)
    scrollbar2 = ttk.Scrollbar(tab2, orient="vertical", command=tree2.yview)
    tree2.configure(yscrollcommand=scrollbar2.set)
    tree2.pack(side="left", fill="both", expand=True)
    scrollbar2.pack(side="right", fill="y")

    # Create a tab containg a treeview filtered to display only Low priority expenses
    tab3 = ttk.Frame(notebook1)
    tree3 = ttk.Treeview(tab3, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree3.column(col, width=90, anchor="w", stretch=tk.YES)
        tree3.heading(col, text=col)
    scrollbar3 = ttk.Scrollbar(tab3, orient="vertical", command=tree3.yview)
    tree3.configure(yscrollcommand=scrollbar3.set)
    tree3.pack(side="left", fill="both", expand=True)
    scrollbar3.pack(side="right", fill="y")

    notebook1.add(tab1, text="High")
    notebook1.add(tab2, text="Medium")
    notebook1.add(tab3, text="Low")
    notebook1.pack(fill="both", expand=True)

    frame4.grid_columnconfigure(0, weight=1)
    frame4.grid_rowconfigure(0, weight=1)

    # Method call to load_expenses() to load any previously entered expenses saved in the excel file into the treeviews
    load_expenses()

    # Frame 3 (Bottom Left)
    # Create two tabs

    if mode == "dark":
        apply_dark_mode_tabs()

    else:
        apply_light_mode_tabs()

    notebook2 = ttk.Notebook(frame3, style="TNotebook")
    tab4 = ttk.Frame(notebook2, style="TNotebook")
    tab5 = ttk.Frame(notebook2, style="TNotebook")

    notebook2.add(tab4, text="Totals")
    notebook2.add(tab5, text="Chart")
    notebook2.pack(fill="both", expand=True)

    frame3.grid_columnconfigure(0, weight=1)
    frame3.grid_rowconfigure(0, weight=1)

    # Tab 4
    # Initialize category labels

    food_label = ctk.CTkLabel(tab4, text="Food: $0", anchor="w", font=("Arial", 28))
    food_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")

    personal_label = ctk.CTkLabel(tab4, text="Personal: $0", anchor="w", font=("Arial", 28))
    personal_label.grid(row=4, column=0, padx=10, pady=5, sticky="w")

    work_label = ctk.CTkLabel(tab4, text="Work: $0", anchor="w", font=("Arial", 28))
    work_label.grid(row=5, column=0, padx=10, pady=5, sticky="w")

    home_label = ctk.CTkLabel(tab4, text="Home: $0", anchor="w", font=("Arial", 28))
    home_label.grid(row=6, column=0, padx=10, pady=5, sticky="w")

    transportation_label = ctk.CTkLabel(tab4, text="Transportation: $0", anchor="w", font=("Arial", 28))
    transportation_label.grid(row=3, column=1, padx=10, pady=5, sticky="w")

    recurring_label = ctk.CTkLabel(tab4, text="Recurring: $0", anchor="w", font=("Arial", 28))
    recurring_label.grid(row=4, column=1, padx=10, pady=5, sticky="w")

    misc_label = ctk.CTkLabel(tab4, text="Miscellaneous: $0", anchor="w", font=("Arial", 28))
    misc_label.grid(row=5, column=1, padx=10, pady=5, sticky="w")


    total_expenses_label = ctk.CTkLabel(frame1, text="Expenses:", anchor="w", font=("Arial", 28))
    total_expenses_label.grid(row=0, column=0, padx=10, pady=20, sticky="w")

    # Initialize funds_remaining_label
    funds_remaining_label = ctk.CTkLabel(frame1, text="${:,.2f}".format(funds_remaining), anchor="w",
                                         font=("Arial", 28))
    funds_remaining_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")

    # Frame 1 (Top Left)
    # Create Add New Expense button
    add_new_expense_button = ctk.CTkButton(frame1, text="Add New Expense", command=expense_popup)
    add_new_expense_button.grid(row=2, column=1, padx=10, pady=10)

    # Create Delete Expense button
    delete_expense_button = ctk.CTkButton(frame1, text="Delete Expense", command=delete_expense)
    delete_expense_button.grid(row=3, column=1, padx=10, pady=10)

    # Create Upload Receipt button
    upload_receipt_button = ctk.CTkButton(frame1, text="Upload Receipt", command=upload_image)
    upload_receipt_button.grid(row=2, column=2, padx=10, pady=10)

    # Create Clear All Expenses button
    clear_all_expenses_button = ctk.CTkButton(frame1, text="Clear All Expenses", command=messagebox_popup)
    clear_all_expenses_button.grid(row=3, column=2, padx=10, pady=10)

    # Method call to update the labels
    #update_labels()

    apply_theme_to_labels_in_tab4(mode)

    # Tab 5
    # Method to call to update_pie_chart to create and update the pie chart
    update_pie_chart()

    welcome_window()
    # Start the main window event loop
    root.mainloop()


# Method to load any previously entered expenses by the user from the excel file into the treeviews
def load_expenses():
    # Load the excel file from the filepath and access the active sheet
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Load the expense info from the active sheet of the excel file into Frame 2 treeview

    # Start the iteration from row 2 of the sheet and yield cell values only
    # (rather than cell objects in order to save memory since we do not need additional cell properties)
    for current_row in sheet1.iter_rows(min_row=2, values_only=True):
        # Parent argument is empty: ""
        # Enter expense data at the end of the Treeview: "end"
        # Insert expense info extracted from current row of the sheet: expense info = current_row
        tree.insert("", "end", values=current_row)

    # Load the expense info from the active sheet of the excel file into Frame 4 treeview tabs
    for column in sheet1.iter_rows(min_row=2, values_only=True):

        specified_columns = (column[0], column[1], column[2], column[4])

        if column[3] == "High":
            tree1.insert("", "end", values=specified_columns)

        elif column[3] == "Medium":
            tree2.insert("", "end", values=specified_columns)

        else:
            tree3.insert("", "end", values=specified_columns)


# Update labels after expense addition
def update_labels(expense_type, price):
    labels = {
        "Food": food_label,
        "Personal": personal_label,
        "Work": work_label,
        "Home": home_label,
        "Transportation": transportation_label,
        "Recurring": recurring_label,
        "Miscellaneous": misc_label
    }

    current_text = labels[expense_type].cget("text")
    current_amount = float(current_text.split("$")[1].replace(",", ""))
    new_amount = current_amount + price
    labels[expense_type].configure(text=f"{expense_type}: ${new_amount:,.2f}")


# Method to update the pie chart
def update_pie_chart():
    # Clear all widgets inside Frame 3 in order to keep the pie chart up to date after an expense is added/deleted
    for widget in tab5.winfo_children():
        widget.destroy()

    # Load existing expenses from the Excel file
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Create a dictionary and set expense Type as keys and expense Price as values initialized at zero
    type_dictionary = {'Food': 0, 'Personal': 0, 'Home': 0, 'Work': 0, 'Transportation': 0, 'Recurring': 0, 'Miscellaneous': 0}

    for row in sheet1.iter_rows(min_row=2, values_only=True):

        expense_type = row[1]
        expense_price = row[2]

        try:
            expense_price = float(expense_price)
        except ValueError:
            continue

        if expense_type in type_dictionary:
            type_dictionary[expense_type] += expense_price

        else:
            type_dictionary[expense_type] = expense_price

    filtered_type_dictionary = {key: value for key, value in type_dictionary.items() if value != 0}

    # Create pie chart labels and sizes for non-zero expense types
    pie_chart_labels = list(filtered_type_dictionary.keys())
    pie_chart_sizes = list(filtered_type_dictionary.values())

    # Create the pie chart in Tab 5 of Frame 3
    fig, ax = plt.subplots()
    legend_font = FontProperties(family='Arial', weight='normal', size=12)
    title_font = FontProperties(family='Arial', weight='bold', size=20)
    ax.pie(pie_chart_sizes, autopct='%1.1f%%', pctdistance=1.25)
    ax.legend(pie_chart_labels, loc="center right", prop=legend_font, bbox_to_anchor=(1.6, 0.5))
    ax.set_title('Expense Distribution by Category', fontproperties=title_font)

    canvas = FigureCanvasTkAgg(fig, master=tab5)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)


# Method to allow the user to input their expense information
def expense_popup():
    global popup, entry1, entry2, entry3, priority_combobox, date_entry

    # Create the expense pop-up window
    popup = ctk.CTkToplevel(root)
    popup.title("Add New Expense")
    popup.geometry("300x450")

    popup.transient(root)
    popup.grab_set()

    # Create labels and entry fields, comboboxes, and calendar for the user to enter their expense information
    label1 = ctk.CTkLabel(popup, text="Expense Name:")
    label1.pack(padx=10, pady=10)
    entry1 = ctk.CTkEntry(popup)
    entry1.pack()

    label2 = ctk.CTkLabel(popup, text="Expense Type:")
    label2.pack(padx=10, pady=10)
    expense_types = ["Food", "Personal", "Work", "Home", "Transportation", "Recurring", "Miscellaneous"]
    entry2 = ctk.CTkComboBox(popup, values=expense_types, button_color="#2FA571")
    entry2.pack()

    label3 = ctk.CTkLabel(popup, text="Expense Price:")
    label3.pack(padx=10, pady=10)
    entry3 = ctk.CTkEntry(popup)
    entry3.pack()

    label4 = ctk.CTkLabel(popup, text="Expense Priority:")
    label4.pack(padx=10, pady=10)
    priority_combobox = ctk.CTkComboBox(popup, values=["High", "Medium", "Low"], button_color="#2FA571")
    priority_combobox.pack()

    label5 = ctk.CTkLabel(popup, text="Expense Date:")
    label5.pack(padx=10, pady=10)
    date_entry = DateEntry(popup, date_pattern="yyyy-mm-dd", background="#2FA571", selectbackground="#2FA571", showweeknumbers=False, firstweekday="sunday", width=25, buttoncolor="#2FA571")
    date_entry.pack()

    add_button = ctk.CTkButton(popup, text="Add", command=add_expense)
    add_button.pack(padx=10, pady=20)


# Method to add the user's expense (modified from "Code First with Hala" YouTube video)
def add_expense():
    try:
        # Validate that all fields are filled
        if not entry1.get() or not entry2.get() or not entry3.get() or not priority_combobox.get() or not date_entry.get():
            messagebox.showerror("Error", "Please fill in all fields")
            return

        # Validate that price is a number
        try:
            price = float(entry3.get())
        except ValueError:
            messagebox.showerror("Error", "Price must be a number")
            return

        if len(entry1.get()) > 15:
            messagebox.showerror("Error", "Expense Name must be 15 characters or less")
            return

        # Append and save the user's expense information to the excel file
        expense_info = [entry1.get(), entry2.get(), price, priority_combobox.get(), date_entry.get()]
        workbook = openpyxl.load_workbook(expense_file)
        sheet1 = workbook.active
        sheet1.append(expense_info)
        workbook.save(expense_file)

        # Insert expense info into Frame 2 treeview
        tree.insert('', tk.END, values=expense_info)

        # Insert expense info into proper Frame 4 treeview tab based on the priority level of the expense
        if priority_combobox.get() == "High":
            tree1.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))
        elif priority_combobox.get() == "Medium":
            tree2.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))
        else:
            tree3.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))

        # Update the labels after expense addition
        #update_labels()

        # Update the pie chart after expense addition
        update_pie_chart()

    except FileNotFoundError:
        messagebox.showerror("Error", f"File {expense_file} not found. Please check the file path and try again.")
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Error", f"The file {expense_file} is not a valid Excel file. Please check the file and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")


# Method to delete selected expense from the treeviews and excel file
def delete_expense():
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Assign selected_item with the currently selected expense row from Frame 2 treeview
    selected_item = tree.selection()[0]
    values = tree.item(selected_item, "values")

    # Delete selected expense from Frame 2 treeview
    tree.delete(selected_item)

    # Delete selected expense from proper Frame 4 treeview tab
    if values[3] == "High":
        for item in tree1.get_children():
            if tree1.item(item, "values")[0] == values[0]:
                tree1.delete(item)
                break
    elif values[3] == "Medium":
        for item in tree2.get_children():
            if tree2.item(item, "values")[0] == values[0]:
                tree2.delete(item)
                break
    elif values[3] == "Low":
        for item in tree3.get_children():
            if tree3.item(item, "values")[0] == values[0]:
                tree3.delete(item)
                break

    # Delete selected expense from Excel file
    for row in sheet1.iter_rows():
        if (row[0].value == values[0] and
                row[1].value == values[1] and
                row[2].value == float(values[2]) and
                row[3].value == values[3] and
                row[4].value == values[4]):
            sheet1.delete_rows(row[0].row, 1)
            break

    workbook.save(expense_file)

    # Update the labels after expense deletion
    #update_labels()

    # Update the pie chart after expense deletion
    update_pie_chart()


# Pop-up window to ask the user to confirm whether or not they want to delete all their entered expenses
def messagebox_popup():
    popup = ctk.CTkToplevel(root)
    popup.title("Confirmation")
    
    popup_width = 350
    popup_height = 250

    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    # Calculate the x and y coordinates for the window to be centered
    x = (screen_width // 2) - (popup_width // 2)
    y = (screen_height // 2) - (popup_height // 2)

    # Set the geometry of the window
    popup.geometry(f'{popup_width}x{popup_height}+{x}+{y}')

    popup.transient(root)
    popup.grab_set()

    messagebox_label1 = ctk.CTkLabel(popup, text=f"Clear All Expenses?", font=("Arial", 16, "bold"))
    messagebox_label1.pack(padx=10, pady=10)
    
    messagebox_label2 = ctk.CTkLabel(popup, text=f"All entered expenses will be permanently deleted.\nThis action cannot be undone.", font=("Arial", 14))
    messagebox_label2.pack(padx=10, pady=10)
    
    confirm_button = ctk.CTkButton(popup, text="Confirm", command=lambda: clear_all_expenses(popup))
    confirm_button.pack(padx=10, pady=10)
    
    cancel_button = ctk.CTkButton(popup, text="Cancel", command=lambda: close_messagebox_popup(popup))
    cancel_button.pack(padx=10, pady=10)
    

# Method to delete all the user's entered expenses from treeviews, excel file, labels, and pie chart
def clear_all_expenses(popup):  
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Delete all entered expenses from all treeviews
    for row in tree.get_children():
        tree.delete(row)
    
    for row in tree1.get_children():
        tree1.delete(row)

    for row in tree2.get_children():
        tree2.delete(row)

    for row in tree3.get_children():
        tree3.delete(row)

    # Delete all entered expenses from excel file
    sheet1.delete_rows(2, sheet1.max_row)
    workbook.save(expense_file)

    # Update the labels after deleting all expenses
    #update_labels()

    # Update the pie chart after deleting all expenses
    update_pie_chart()

    # Method call to close the messagebox pop-up window
    close_messagebox_popup(popup)


# Method to close the messsagebox pop-up window
def close_messagebox_popup(popup):
    popup.destroy()


# Normalize keys in the extracted receipt data
def normalize_keys(data):
    if isinstance(data, dict):
        return {key.lower().replace(' ', '_'): normalize_keys(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [normalize_keys(element) for element in data]
    return data


# Extract data from the receipt image using OCR and OpenAI
def extract_receipt_data(file_path):
    try:
        text = pytesseract.image_to_string(Image.open(file_path))
        prompt = f"Extract the relevant data from the following receipt text and provide it in JSON format:\n\n{text}"
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            n=1,
            stop=None,
            temperature=0.5
        )

        parsed_text = response.choices[0].message.content.strip().strip('```json').strip('```')

        try:
            receipt_data = json.loads(parsed_text)
            normalized_data = normalize_keys(receipt_data)
            return normalized_data
        except json.JSONDecodeError as json_err:
            print(f"JSON decode error: {json_err}")
            print(f"Original parsed text: {parsed_text}")
            return None
    except Exception as e:
        print(f"Failed to extract data from receipt: {e}")
        return None


# Upload and process receipt image
def upload_image():

    file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
    if file_path:
        receipt_data = extract_receipt_data(file_path)
        if receipt_data:

            print("Extracted Receipt Data:", json.dumps(receipt_data, indent=4))

            store_name = (
                    receipt_data.get('store_name') or
                    receipt_data.get('store') or
                    receipt_data.get('store_info', {}).get('store_name') or
                    receipt_data.get('store_information', {}).get('store_name') or
                    receipt_data.get('location', {}).get('store_name') or
                    receipt_data.get('storeinfo', {}).get('storename') or
                    receipt_data.get('store', {}).get('name') or
                    receipt_data.get('storeinfo', {}).get('store') or
                    receipt_data.get('merchant') or
                    receipt_data.get('store_info', {}).get('name') or
                    'Unknown Store'
            )
            total_price = str(
                receipt_data.get('total_balance') or
                receipt_data.get('total') or
                receipt_data.get('balance_due') or
                receipt_data.get('totals', {}).get('total') or
                receipt_data.get('payment', {}).get('total_amount') or
                receipt_data.get('total_purchase') or
                receipt_data.get('receipt', {}).get('total') or
                receipt_data.get('balance') or
                receipt_data.get('payment_info', {}).get('total_purchase') or
                receipt_data.get('amount_due') or
                receipt_data.get('total_amount') or
                receipt_data.get('receiptinfo', {}).get('total') or
                'Unknown Price'
            )
            date = (
                    receipt_data.get('transaction_date') or
                    receipt_data.get('date') or
                    receipt_data.get('date_time') or
                    receipt_data.get('transaction_info', {}).get('transaction_date') or
                    receipt_data.get('transaction_details', {}).get('date') or
                    receipt_data.get('payment_information', {}).get('transaction_date') or
                    receipt_data.get('receipt', {}).get('date') or
                    receipt_data.get('transaction_date_time') or
                    receipt_data.get('TransactionDateTime') or
                    receipt_data.get('purchase_date') or
                    receipt_data.get('receiptinfo', {}).get('date') or
                    receipt_data.get('receipt_date') or
                    receipt_data.get('transaction_details', {}).get('date_time') or
                    receipt_data.get('store', {}).get('date') or
                    'Unknown Date'
            )

            # Add pop-up inputs for expense type and priority
            expense_type = simpledialog.askstring("Input",
                                                  "Enter the expense type (Food, Personal, Work, Home, Transportation, Recurring, Miscellaneous):",
                                                  parent=root)
            priority = simpledialog.askstring("Input", "Enter the priority (High, Medium, Low):", parent=root)

            if not total_price or total_price == 'Unknown Price':
                messagebox.showerror("Error", "Total price is required.")
                return

            if not expense_type:
                messagebox.showerror("Error", "Expense type is required.")
                return

            if not priority:
                messagebox.showerror("Error", "Priority is required.")
                return

            add_expense_from_receipt(store_name, total_price, date, expense_type, priority)
        else:
            messagebox.showerror("Error", "Could not extract data from the receipt.")


# Modify the add_expense_from_receipt function to accept expense_type and priority parameters
def add_expense_from_receipt(store_name, total_price, date, expense_type, priority):
    global funds_remaining  # Declare funds_remaining as global

    try:
        price = float(total_price)
    except ValueError:
        messagebox.showerror("Error", "Invalid price format.")
        return

    funds_remaining -= price
    funds_remaining_label.configure(text="${:,.2f}".format(funds_remaining))  # Use 'configure' instead of 'config'

    tree.insert("", "end", values=(store_name, expense_type, price, priority, date))

    if priority == "High":
        tree1.insert("", "end", values=(store_name, expense_type, price, date))
    elif priority == "Medium":
        tree2.insert("", "end", values=(store_name, expense_type, price, date))
    else:
        tree3.insert("", "end", values=(store_name, expense_type, price, date))

    update_labels(expense_type, price)
    save_to_excel(store_name, expense_type, price, priority, date)


# Save to excel file
def save_to_excel(name, expense_type, price, priority, date):
    if os.path.exists(expense_file):
        df = pd.read_excel(expense_file)
    else:
        df = pd.DataFrame(columns=["Name", "Type", "Price", "Priority", "Date"])

    new_row = {"Name": name, "Type": expense_type, "Price": price, "Priority": priority, "Date": date}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(expense_file, index=False)


# Method to apply the light mode to the app
def apply_light_mode_style():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("Treeview",
                    background="#FFFFFF",
                    foreground="black",
                    rowheight=25,
                    fieldbackground="#FFFFFF")

    style.map('Treeview', background=[('selected', '#E5E5E5')])

    style.configure("Treeview.Heading",
                    background="#DDDDDD",
                    foreground="black",
                    font=("Helvetica", 10, "bold"))

def apply_dark_mode_style():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("Treeview",
                    background="#333333",
                    foreground="white",
                    rowheight=25,
                    fieldbackground="#333333")

    style.map('Treeview', background=[('selected', '#2b2b2b')])

    style.configure("Treeview.Heading",
                    background="#444444",
                    foreground="white",
                    font=("Helvetica", 10, "bold"))

def apply_dark_mode_style_welcome():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("TLabel",
                    background="#333333",  # Dark background
                    foreground="white")  # White text

def apply_light_mode_style_welcome():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("TLabel",
                    background="#FFFFFF",  # Light background
                    foreground="black")  # Black text

def apply_dark_mode_tabs():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("TNotebook", background="#333333")

def apply_light_mode_tabs():
    style = ttk.Style()
    style.theme_use("default")

    style.configure("TNotebook", background="#FFFFFF")

def apply_theme_to_labels_in_tab4(mode):
    global food_label, personal_label, work_label, home_label, transportation_label, recurring_label, misc_label

    # Create a list of labels in tab4
    labels = [
        food_label,
        personal_label,
        work_label,
        home_label,
        transportation_label,
        recurring_label,
        misc_label
    ]

    if mode == "dark":
        fg_color = "#333333"
    else:
        fg_color = "#FFFFFF"

    # Determine foreground color based on the theme
    for label in labels:
        label.configure(fg_color=fg_color)


# Main window call to start the program
if __name__ == '__main__':
    main_window()
