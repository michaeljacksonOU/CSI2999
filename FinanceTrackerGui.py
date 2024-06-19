import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import pandas as pd
import os
from tkcalendar import DateEntry
import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Note: assignment of the file path must be changed accordingly to the unique location where is saved on each individual user's end system 
expense_file = r"C:\Users\Peter\PycharmProjects\pythonProject\expenses.xlsx"

# Set app mode and theme
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
mode = "dark"

# Method to switch the app mode between light and dark mode
def change_theme():
    global mode
    if mode == "dark":
        ctk.set_appearance_mode("light")
        mode = "light"
        apply_light_mode_style()
    else:
        ctk.set_appearance_mode("dark")
        mode = "dark"
        apply_dark_mode_style()


# Method to display main window elements:
# Frame 1: budget and funds remaining
# Frame 2: treeview of all the user's entered expenses
# Frame 3: pie chart of distribution of entered expenses by category
# Frame 4: treeview of entered expenses sorted by priority tabs 
def main_window():
    global root, tree, tree1, tree2, tree3

    # Create the main window 
    root = ctk.CTk()
    root.title("ExpenseExpert")
    root.geometry("1280x720")

    # Create and place the frames in the grid
    frame1 = ctk.CTkFrame(root)
    frame2 = ctk.CTkFrame(root)
    frame3 = ctk.CTkFrame(root)
    frame4 = ctk.CTkFrame(root)
    frame5 = ctk.CTkFrame(root)
    frame1.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    frame2.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    frame3.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
    frame4.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
    frame5.grid(row=2, column=0, padx=10, pady=10, sticky="w")

    # Set propagation to false for all frames in order to prevent them from resizing based on the size of the children widgets within them
    frame1.propagate(False)
    frame2.propagate(False)
    frame3.propagate(False)
    frame4.propagate(False)
    frame5.propagate(False)

    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    # Create a toggle bar to allow the user to switch between light and dark mode
    switch_var = ctk.StringVar(value="on")
    switch = ctk.CTkSwitch(frame5, text="Mode", command=change_theme, variable=switch_var, onvalue="on", offvalue="off")
    switch.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    label1 = ctk.CTkLabel(frame1, text="Expenses:", anchor="w", font=("Arial", 28))
    label1.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    #expenses_amount_label = ctk.CTkLabel(frame1, text="${:,.2f}".format(budget_amount), anchor="w", font=("Arial", 28))
    #expenses_amount_label.grid(row=0, column=1, padx=10, pady=10, sticky="w")

    # Frame 1 (Top Left)
    # Create Add New Expense button
    add_new_expense_button = ctk.CTkButton(frame1, text="Add New Expense", command=lambda: expense_popup(frame3))
    add_new_expense_button.grid(row=2, column=0, padx=10, pady=10)

    # Create Delete Expense button
    delete_expense_button = ctk.CTkButton(frame1, text="Delete Expense", command=lambda: delete_expense(frame3))
    delete_expense_button.grid(row=2, column=1, padx=10, pady=10)

    # Frame 2 (Top Right)
    if mode == "dark":
        apply_dark_mode_style()
    else:
        apply_light_mode_style()

    # Create a treeview to display the user's entered expenses
    tree = ttk.Treeview(frame2, show="headings", style="Treeview")
    tree["columns"] = ("Name", "Type", "Price", "Priority", "Date")

    for col in tree["columns"]:
        tree.column(col, width=100, anchor="w", stretch="YES")
        tree.heading(col, text=col, anchor="w")

    scrollbar = ttk.Scrollbar(frame2, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=1, sticky='ns')

    tree.configure(yscrollcommand=scrollbar.set)
    tree.grid(row=0, column=0, sticky="nsew")

    frame2.grid_rowconfigure(0, weight=1)
    frame2.grid_columnconfigure(0, weight=1)

    # Frame 4 (Bottom Right)
    style = ttk.Style()
    style.configure("TNotebook.Tab", font=('Helvetica', 25, 'bold'), padding=[40, 0])
    notebook = ttk.Notebook(frame4, style="TNotebook")

    # Create a tab containing a treeview filtered to display only High priority expenses
    tab1 = ttk.Frame(notebook)
    tree1 = ttk.Treeview(tab1, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree1.column(col, width=90, anchor="w", stretch=tk.YES)
        tree1.heading(col, text=col)
    scrollbar1 = ttk.Scrollbar(tab1, orient="vertical", command=tree1.yview)
    tree1.configure(yscrollcommand=scrollbar1.set)
    tree1.pack(side="left", fill="both", expand=True)
    scrollbar1.pack(side="right", fill="y")

    # Create a tab containing a treeview filtered to display only Medium priority expenses
    tab2 = ttk.Frame(notebook)
    tree2 = ttk.Treeview(tab2, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree2.column(col, width=90, anchor="w", stretch=tk.YES)
        tree2.heading(col, text=col)
    scrollbar2 = ttk.Scrollbar(tab2, orient="vertical", command=tree2.yview)
    tree2.configure(yscrollcommand=scrollbar2.set)
    tree2.pack(side="left", fill="both", expand=True)
    scrollbar2.pack(side="right", fill="y")

    # Create a tab containing a treeview filtered to display only Low priority expenses
    tab3 = ttk.Frame(notebook)
    tree3 = ttk.Treeview(tab3, columns=("Name", "Type", "Price", "Date"), show="headings", style="Treeview")
    for col in ("Name", "Type", "Price", "Date"):
        tree3.column(col, width=90, anchor="w", stretch=tk.YES)
        tree3.heading(col, text=col)
    scrollbar3 = ttk.Scrollbar(tab3, orient="vertical", command=tree3.yview)
    tree3.configure(yscrollcommand=scrollbar3.set)
    tree3.pack(side="left", fill="both", expand=True)
    scrollbar3.pack(side="right", fill="y")

    notebook.add(tab1, text="High")
    notebook.add(tab2, text="Medium")
    notebook.add(tab3, text="Low")
    notebook.pack(fill="both", expand=True)

    frame4.grid_columnconfigure(0, weight=1)
    frame4.grid_rowconfigure(0, weight=1)
    
    # Method call to load_expenses() to load any previously entered expenses saved in the excel file into the treeviews 
    load_expenses()

    # Frame 3 (Bottom LeftS)
    # Method to call to update_pie_chart to create and update the pie chart
    update_pie_chart(frame3)

    # Start the main window event loop
    root.mainloop()


# Method to load any previously entered expenses by the user from the excel file into the treeviews
def load_expenses():
    # Load the excel file from the filepath and access the active sheet
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Load the expense info from the active sheet of the excel file into Frame 2 treeview

    # Start the iteration from row 2 of the sheet and yield cell values only
    # (rather than cell objects in order to save memory since we do not need additional cell properties)
    for current_row in sheet1.iter_rows(min_row=2, values_only=True):
        # Parent argument is empty: ""
        # Enter expense data at the end of the Treeview: "end"
        # Insert expense info extracted from current row of the sheet: expense info = current_row
        tree.insert("", "end", values=current_row)

    # Load the expense info from the active sheet of the excel file into Frame 4 treeview tabs
    for column in sheet1.iter_rows(min_row=2, values_only=True):

        specified_columns = (column[0], column[1], column[2], column[4])

        if column[3] == "High":
            tree1.insert("", "end", values=specified_columns)

        elif column[3] == "Medium":
            tree2.insert("", "end", values=specified_columns)

        else:
            tree3.insert("", "end", values=specified_columns)


def update_pie_chart(frame3):
    # Clear all widgets inside Frame 3 in order to keep the pie chart up to date after an expense is added/deleted
    for widget in frame3.winfo_children():
        widget.destroy()

    # Load existing expenses from the Excel file
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Create a dictionary and set expense Type set as keys and expense Price as values initialized at zero
    type_dictionary = {'Food': 0, 'Personal': 0, 'Home': 0, 'Work': 0, 'Transportation': 0, 'Recurring': 0, 'Miscellaneous': 0}

    for row in sheet1.iter_rows(min_row=2, values_only=True):

        expense_type = row[1]
        expense_price = row[2]

        try:
            expense_price = float(expense_price)
        except ValueError:
            continue

        if expense_type in type_dictionary:
            type_dictionary[expense_type] += expense_price

        else:
            type_dictionary[expense_type] = expense_price

    filtered_type_dictionary = {key: value for key, value in type_dictionary.items() if value != 0}

    # Create pie chart labels and sizes for non-zero expense types
    pie_chart_labels = list(filtered_type_dictionary.keys())
    pie_chart_sizes = list(filtered_type_dictionary.values())

    # Create the pie chart in Frame 3
    fig, ax = plt.subplots()
    ax.pie(pie_chart_sizes, autopct='%1.1f%%', pctdistance=1.25)
    ax.legend(pie_chart_labels, loc="center right", bbox_to_anchor=(1.75, 0.5))
    ax.set_title('Expense Distribution by Category')

    canvas = FigureCanvasTkAgg(fig, master=frame3)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)


# Method to allow the user to input their expense information
def expense_popup(frame3):
    global popup, entry1, entry2, entry3, priority_combobox, date_entry

    # Create the expense pop-up window
    popup = ctk.CTkToplevel(root)
    popup.title("Add New Expense")
    popup.geometry("300x300")

    popup.transient(root)
    popup.grab_set()

    # Create labels and entry fields, comboboxes, and calendar for the user to enter their expense information
    label1 = ctk.CTkLabel(popup, text="Expense Name:")
    label1.pack()
    entry1 = ctk.CTkEntry(popup)
    entry1.pack()

    label2 = ctk.CTkLabel(popup, text="Expense Type:")
    label2.pack()
    expense_types = ["Food", "Personal", "Work", "Home", "Transportation", "Recurring", "Miscellaneous"]
    entry2 = ctk.CTkComboBox(popup, values=expense_types)
    entry2.pack()

    label3 = ctk.CTkLabel(popup, text="Expense Price:")
    label3.pack()
    entry3 = ctk.CTkEntry(popup)
    entry3.pack()

    label4 = ctk.CTkLabel(popup, text="Expense Priority:")
    label4.pack()
    priority_combobox = ctk.CTkComboBox(popup, values=["High", "Medium", "Low"])
    priority_combobox.pack()

    label5 = ctk.CTkLabel(popup, text="Expense Date:")
    label5.pack()
    date_entry = DateEntry(popup, date_pattern="yyyy-mm-dd")
    date_entry.pack()

    add_button = tk.Button(popup, text="Add", command=lambda: add_expense(frame3))
    add_button.pack()

# Method add_expense() modified from "Code First with Hala" YouTube video:
# https://www.youtube.com/watch?v=8m4uDS_nyCk&pp=ygUlcHl0aG9uIGV4Y2VsIGFwcCBjb2RlIGZpcnN0IHdpdGggaGFsYQ%3D%3D
def add_expense(frame3):
    try:
        # Validate that all fields are filled
        if not entry1.get() or not entry2.get() or not entry3.get() or not priority_combobox.get() or not date_entry.get():
            messagebox.showerror("Error", "Please fill in all fields")
            return

        # Validate that price is a number
        try:
            price = float(entry3.get())
        except ValueError:
            messagebox.showerror("Error", "Price must be a number")
            return

        if len(entry1.get()) > 15:
            messagebox.showerror("Error", "Expense Name must be 15 characters or less")
            return

        # Append and save the user's expense information to the excel file
        expense_info = [entry1.get(), entry2.get(), price, priority_combobox.get(), date_entry.get()]
        workbook = openpyxl.load_workbook(expense_file)
        sheet1 = workbook.active
        sheet1.append(expense_info)
        workbook.save(expense_file)

        # Insert expense info into Frame 2 treeview
        tree.insert('', tk.END, values=expense_info)

        # Insert expense info into proper Frame 4 treeview tab based on the priority level of the expense
        if priority_combobox.get() == "High":
            tree1.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))
        elif priority_combobox.get() == "Medium":
            tree2.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))
        else:
            tree3.insert("", "end", values=(entry1.get(), entry2.get(), price, date_entry.get()))

        # Update the pie chart
        update_pie_chart(frame3)
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {expense_file} not found. Please check the file path and try again.")
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Error", f"The file {expense_file} is not a valid Excel file. Please check the file and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

# Method to delete selected expense from the treeviews and excel file
def delete_expense(frame3):
    workbook = openpyxl.load_workbook(expense_file)
    sheet1 = workbook.active

    # Assign selected_item with the currently selected expense row from Frame 2 treeview
    selected_item = tree.selection()[0]
    values = tree.item(selected_item, "values")

    # Delete selected expense from Frame 2 treeview
    tree.delete(selected_item)

    # Delete selected expense from proper Frame 4 treeview tab
    if values[3] == "High":
        for item in tree1.get_children():
            if tree1.item(item, "values")[0] == values[0]:
                tree1.delete(item)
                break
    elif values[3] == "Medium":
        for item in tree2.get_children():
            if tree2.item(item, "values")[0] == values[0]:
                tree2.delete(item)
                break
    elif values[3] == "Low":
        for item in tree3.get_children():
            if tree3.item(item, "values")[0] == values[0]:
                tree3.delete(item)
                break

    # Delete selected expense from Excel file
    for row in sheet1.iter_rows():
        if (row[0].value == values[0] and
                row[1].value == values[1] and
                row[2].value == float(values[2]) and
                row[3].value == values[3] and
                row[4].value == values[4]):
            sheet1.delete_rows(row[0].row, 1)
            break

    workbook.save(expense_file)

    # Method call to update_pie_chart() to update the pie chart after expense deletion
    update_pie_chart(frame3)

def apply_light_mode_style():
    style = ttk.Style()
    style.theme_use("alt")

    style.configure("Treeview",
                    background="#FFFFFF",  # Light background
                    foreground="black",  # Black text
                    rowheight=25,
                    fieldbackground="#FFFFFF")  # Light background

    style.map('Treeview', background=[('selected', '#E5E5E5')])  # Selected row color

    style.configure("Treeview.Heading",
                    background="#DDDDDD",  # Light background for headings
                    foreground="black",  # Black text for headings
                    font=("Helvetica", 10, "bold"))  # Bold font for headings

def apply_dark_mode_style():
    style = ttk.Style()
    style.theme_use("alt")

    style.configure("Treeview",
                    background="#333333",  # Dark background
                    foreground="white",  # White text
                    rowheight=25,
                    fieldbackground="#333333")  # Dark background

    style.map('Treeview', background=[('selected', '#2b2b2b')])  # Selected row color

    style.configure("Treeview.Heading",
                    background="#444444",  # Darker background for headings
                    foreground="white",  # White text for headings
                    font=("Helvetica", 10, "bold"))  # Bold font for headings


# Method call to main_window() to start the program 
main_window()
